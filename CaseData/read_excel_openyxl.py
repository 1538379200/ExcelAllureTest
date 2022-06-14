import openpyxl
import pytest
from BaseControl.base import BaseControl
import allure
from openpyxl.styles import PatternFill
from openpyxl.formatting.formatting import ConditionalFormattingList
from pathlib import Path
import os
"""
openpyxl的形式，添加了excel填入的形式
当表格中设置了assert的方法的时候，代码将在reports文件夹下生成新的TestResult.xlsx文件，区别于用例文件
assert断言失败的用例将以红色前景色作为标识
"""

keys = None                 # 全局变量key，用来接受BaseControl实例
old_model = None            # 全局变量old_model，用来保存旧数据的模块名称
old_story = None            # 全局变量old_story，用来保存旧的story(流程名称)
report_path = Path(__file__).resolve().parents[1]/'reports'

def get_cases():
    workbook = openpyxl.load_workbook('NewExcelData.xlsx')
    sheets = workbook.sheetnames                # 获取所有sheet名称
    sys_name = '未获取到的默认值'
    all_cases = []
    sheet_cases = {}
    for sheet in sheets:
        data_sheet = workbook[sheet]            # 获取sheet内容
        case_data = list(data_sheet.values)     # 获取sheet所有数据，并强转为list格式
        if case_data[0][0] is not None:
            sys_name = case_data[0][0]          # 获取第一行为用例系统名称
        test_cases = case_data[2:]              # 获取除1、2行的数据，作为测试所需用例
        sheet_cases[sheet] = test_cases         # 将所有数据转换为sheet:value的字典形式(后面操作为写入excel做准备)
    for val in list(sheet_cases.items()):       # 循环所有sheet
        for case in val[1]:                     # 取出所有用例数据
            case_list = [val[0], *case]         # 拼接sheet和用例，组合为一个list
            all_cases.append(case_list)         # 全部存入一个大list作为返回值
    return sys_name, all_cases, workbook      # 返回系统名称、用例表、openpyxl工作文本

class TestExcelYxl:
    sys_name, data_list, workbook = get_cases()

    @allure.epic(sys_name)
    @pytest.mark.parametrize('sheet_name, case_id, model, process, event, location, elements, txt, desc', data_list)
    def test_cases(self, sheet_name, case_id, model, process, event, location, elements, txt, desc):
        global keys, old_story, old_model   # 指定全局变量
        if model is None:                   # 当获取的数据为nan时，使用旧数据作为模块名
            case_model = old_model
        else:
            case_model = model
        allure.dynamic.feature(case_model)  # 动态设置allure的模块名称
        allure.dynamic.title(desc)          # 动态设置用例的标题
        if process is None:                 # 当流程为nan时，设置story为旧数据
            case_story = old_story
        else:
            case_story = process
        allure.dynamic.story(case_story)    # 动态设置allure的story
        action_data = [x for x in [location, elements, txt] if x is not None]  # 过滤为nan的空用例数据
        if event == "open_browser":
            keys = BaseControl(txt)
        else:
            with allure.step(desc):         # 设置用例的step，以表格的描述为准
                assert_res = getattr(keys, 'event_' + event)(*action_data)
                if event == 'assert':                           # 当用户数据中有assert要求时进行下列写入操作
                    sheet = self.workbook[sheet_name]
                    sheet.conditional_formatting = ConditionalFormattingList()  # 清除原表格条件格式
                    if assert_res is False:             # 断言失败时
                        for col in range(1, 9):
                            # 将红色填充为前景色
                            sheet.cell(row=case_id+2, column=col).fill = PatternFill(fgColor='E74A4A', fill_type='solid')
                        raise AssertionError(f"断言失败，用户数据为：{txt}")       # 将此用例结果作为失败用例展现，避免表格报告结果不一致
                    self.workbook.save(report_path / 'TestResult.xlsx')
        old_story = case_story              # 设置旧的story
        old_model = case_model              # 设置旧的model


if __name__ == '__main__':
    result_path = report_path/'results'
    pytest.main(['-s', '-v', 'read_excel_openyxl.py', '--alluredir', '../reports/results', '--clean-alluredir'])
    os.system(f'allure serve {result_path}')
